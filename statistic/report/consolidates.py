from copy import deepcopy

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


def consolidate_base(ws1, d1, d2, fin_source):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = f'Сводный: {fin_source}'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'

    columns = [
        ('Номер карты', 10),
        ('Возраст', 8),
        ('Фамилия', 16),
        ('Имя', 16),
        ('Отчество', 16),
        ('Комментарий_в_карте', 20),
        ('Наимен. услуги', 30),
        ('№ напраления', 15),
        ('Дата подтверждения', 15),
        ('Врач', 25),
        ('Место работы', 55),
        ('Профиль_мед.помощи', 15),
        ('Спец_врача', 25),
        ('Основное', 15),
        ('Подчинение', 15),
        ('Цель', 15),
        ('Категория', 25),
    ]

    row = 5
    for idx, column in enumerate(columns, 1):
        ws1.cell(row=row, column=idx).value = column[0]
        ws1.column_dimensions[get_column_letter(idx)].width = column[1]
        ws1.cell(row=row, column=idx).style = style_border

    return ws1


def consolidate_fill_data(ws1, result_query):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    r = 5
    tmp_data = {}
    for i in result_query:
        if i.dir_id:
            tmp_data[i.id_iss] = {
                "patient_card_num": i.patient_card_num,
                "patient_age": i.patient_age,
                "patient_family": i.patient_family,
                "patient_name": i.patient_name,
                "patient_patronymic": i.patient_patronymic,
                "dir_harmful_factor": i.dir_harmful_factor,
                "research_title": i.research_title,
                "dir_id": i.dir_id,
                "date_confirm": i.date_confirm,
                "fio": f"{i.doc_f} {i.doc_n} {i.doc_p}",
                "patient_workplace": i.patient_workplace,
                "doc_speciality": i.doc_speciality,
                "purpose": i.purpose,
            }
            r += 1
            ws1.cell(row=r, column=1).value = i.patient_card_num
            ws1.cell(row=r, column=2).value = i.patient_age
            ws1.cell(row=r, column=3).value = i.patient_family
            ws1.cell(row=r, column=4).value = i.patient_name
            ws1.cell(row=r, column=5).value = i.patient_patronymic
            ws1.cell(row=r, column=6).value = i.dir_harmful_factor
            ws1.cell(row=r, column=7).value = i.research_title
            ws1.cell(row=r, column=8).value = i.dir_id
            ws1.cell(row=r, column=9).value = i.date_confirm
            ws1.cell(row=r, column=10).value = f"{i.doc_f} {i.doc_n} {i.doc_p}"
            ws1.cell(row=r, column=11).value = i.patient_workplace
            ws1.cell(row=r, column=12).value = ""
            ws1.cell(row=r, column=13).value = i.doc_speciality
            ws1.cell(row=r, column=14).value = i.id_iss
            ws1.cell(row=r, column=15).value = i.parent_iss
            ws1.cell(row=r, column=16).value = i.purpose
            ws1.cell(row=r, column=17).value = i.category_title
            for j in range(1, 18):
                ws1.cell(row=r, column=j).style = style_border1

        if i.parent_iss:
            data = tmp_data.get(i.parent_iss)
            if not data:
                continue
            r += 1
            ws1.cell(row=r, column=1).value = data["patient_card_num"]
            ws1.cell(row=r, column=2).value = data["patient_age"]
            ws1.cell(row=r, column=3).value = data["patient_family"]
            ws1.cell(row=r, column=4).value = data["patient_name"]
            ws1.cell(row=r, column=5).value = data["patient_patronymic"]
            ws1.cell(row=r, column=6).value = data["dir_harmful_factor"]
            ws1.cell(row=r, column=7).value = i.research_title
            ws1.cell(row=r, column=8).value = data["dir_id"]
            ws1.cell(row=r, column=9).value = data["date_confirm"]
            ws1.cell(row=r, column=10).value = data["fio"]
            ws1.cell(row=r, column=11).value = data["patient_workplace"]
            ws1.cell(row=r, column=12).value = ""
            ws1.cell(row=r, column=13).value = data["doc_speciality"]
            ws1.cell(row=r, column=14).value = i.id_iss
            ws1.cell(row=r, column=15).value = i.parent_iss
            ws1.cell(row=r, column=16).value = data["purpose"]
            ws1.cell(row=r, column=17).value = i.category_title
            for j in range(1, 18):
                ws1.cell(row=r, column=j).style = style_border1

    return ws1


def consolidate_research_sets_base(ws1, d1, d2, fin_source, head_data, company_title, coast_data):
    style_border = NamedStyle(name="style_border_ca5")
    bd = Side(style='thin', color="000000")
    style_border.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border.font = Font(bold=True, size=11)
    style_border.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws1.cell(row=1, column=1).value = f'Сводный: {fin_source}'
    ws1.cell(row=2, column=1).value = 'Период:'
    ws1.cell(row=3, column=1).value = f'c {d1} по {d2}'
    ws1.cell(row=4, column=1).value = f'Контрагент: {company_title}'

    columns = [
        ('Отдел', 20),
        ('№ п/п.', 20),
        ('Карта', 15),
        ('ФИО', 48),
    ]
    start_column_research = len(columns) + 1

    custom_columns = [(i, 13) for i in head_data.values()]
    coast_data_column = [*[("", 10) for k in range(len(columns))], *[(i, 13) for i in coast_data.values()]]
    columns.extend(custom_columns)

    row = 6
    step = 0
    for k, j in zip(columns, coast_data_column):
        step += 1
        ws1.cell(row=row, column=step).value = k[0]
        ws1.cell(row=row + 1, column=step).value = j[0]
        ws1.column_dimensions[get_column_letter(step)].width = k[1]
        ws1.cell(row=row, column=step).style = style_border
        ws1.cell(row=row + 1, column=step).style = style_border
    return (
        ws1,
        start_column_research,
    )


def consolidate_research_sets_fill_data(ws1, query, def_value_data, start_research_column):
    style_border1 = NamedStyle(name="style_border1")
    bd = Side(style='thin', color="000000")
    style_border1.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    style_border1.font = Font(bold=False, size=11)
    style_border1.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    last_patient = None
    base_step = 0
    current_patient_researh_data = deepcopy(def_value_data)
    current_department_id = None
    current_department_title = ""
    last_patient_fio = ""
    last_patient_card = ""
    row = 7
    price_row = row
    start_row = 0
    current_sum_columns = {}
    total_sum_rows = []
    fill_rows = {}
    count_patient_in_department = 0
    for i in query:
        if last_patient != i.client_id and base_step != 0:
            row += 1
            count_patient_in_department += 1
            ws1.cell(row=row, column=1).value = current_department_title
            ws1.cell(row=row, column=2).value = count_patient_in_department
            ws1.cell(row=row, column=3).value = last_patient_card
            ws1.cell(row=row, column=4).value = last_patient_fio
            column = start_research_column
            if start_row == 0:
                start_row = row
            for k in current_patient_researh_data.values():
                ws1.cell(row=row, column=column).value = k
                current_sum_columns[column] = f'=SUM({get_column_letter(column)}{start_row}:{get_column_letter(column)}{row})'
                column += 1

            current_patient_researh_data = deepcopy(def_value_data)

        current_patient_researh_data[i.research_id] = 1

        if current_department_id != i.department_id and base_step != 0:
            count_patient_in_department = 0
            start_row = 0
            row += 1
            ws1.cell(row=row, column=1).value = f'Итого кол-во {current_department_title}'
            for col, val in current_sum_columns.items():
                ws1.cell(row=row, column=col).value = val
                ws1.cell(row=row + 1, column=col).value = f'={get_column_letter(col)}{price_row}*{get_column_letter(col)}{row}'
            total_sum_rows.append(row)
            row += 1
            ws1.cell(row=row, column=1).value = f'Итого сумма {current_department_title}'
            fill_rows[row] = col
            current_sum_columns = {}
        current_department_id = i.department_id
        current_department_title = i.department_title
        last_patient = i.client_id
        last_patient_fio = f"{i.patient_family} {i.patient_name} {i.patient_patronymic}"
        last_patient_card = i.patient_card_num
        base_step += 1

    row += 1
    count_patient_in_department += 1
    ws1.cell(row=row, column=1).value = current_department_title
    ws1.cell(row=row, column=2).value = count_patient_in_department
    ws1.cell(row=row, column=3).value = last_patient_card
    ws1.cell(row=row, column=4).value = last_patient_fio
    column = start_research_column
    for k in current_patient_researh_data.values():
        if start_row == 0:
            start_row = row
        ws1.cell(row=row, column=column).value = k
        current_sum_columns[column] = f'=SUM({get_column_letter(column)}{start_row}:{get_column_letter(column)}{row})'
        column += 1

    row += 1
    ws1.cell(row=row, column=1).value = f'Итого кол-во {current_department_title}'
    for col, val in current_sum_columns.items():
        ws1.cell(row=row, column=col).value = val
        ws1.cell(row=row + 1, column=col).value = f'={get_column_letter(col)}{price_row}*{get_column_letter(col)}{row}'
    total_sum_rows.append(row)
    row += 1
    ws1.cell(row=row, column=1).value = f'Итого сумма {current_department_title}'
    fill_rows[row] = col

    row += 1
    ws1.cell(row=row, column=1).value = 'Итого кол-во всего'
    total_sum_end = ""
    for total_col in range(start_research_column, col + 1):
        step = 0
        for total_research in total_sum_rows:
            if step != 0:
                total_sum_end = f'{total_sum_end} + {get_column_letter(total_col)}{total_research}'
            else:
                total_sum_end = f'{get_column_letter(total_col)}{total_research}'
            step += 1
        ws1.cell(row=row, column=total_col).value = f"={total_sum_end}"
        total_sum_end = ""

    row += 1
    ws1.cell(row=row, column=1).value = 'Итого сумма всего'
    for total_col in range(start_research_column, col + 1):
        ws1.cell(row=row, column=total_col).value = f'={get_column_letter(total_col)}{price_row}*{get_column_letter(total_col)}{row-1}'

    for m in range(row - base_step, row + 1):
        for n in range(1, len(def_value_data) + start_research_column):
            ws1.cell(row=m, column=n).style = style_border1

    total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', start_color='ffcc66', end_color='ffcc66')
    for k, v in fill_rows.items():
        fill_cells(ws1[f'A{k}:{get_column_letter(v)}{k}'], total_fill)
        fill_cells(ws1[f'A{k - 1}:{get_column_letter(v)}{k - 1}'], total_fill)

    return ws1


def fill_cells(rows_fill, total_fill):
    for row_f in rows_fill:
        for cell in row_f:
            cell.fill = total_fill
