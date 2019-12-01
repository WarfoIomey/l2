from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from directory.models import DispensaryRouteSheet


class Command(BaseCommand):
    def add_arguments(self, parser):
        """
        :param path - файл с кодами МКБ10.2019 + расшифровка
        """
        parser.add_argument('path', type=str)


    def handle(self, *args, **kwargs):
        fp = kwargs["path"]
        self.stdout.write("Path: " + fp)
        wb = load_workbook(filename=fp)
        ws = wb[wb.sheetnames[0]]
        starts = False
        for row in ws.rows:
            cells = [str(x.value) for x in row]
            if not starts:
                if "код" in cells:
                    starts = True
                    code = cells.index("код")
            else:
               pass